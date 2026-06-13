import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURATION & PALETTE ---
FILE_NAME = "Chapter2_Drip_Ads_Audit_and_ROAS_Strategy.xlsx"

# Fonts
FONT_NAME = "Arial"
font_title = Font(name=FONT_NAME, size=16, bold=True, color="1E293B")
font_subtitle = Font(name=FONT_NAME, size=11, italic=True, color="64748B")
font_section_header = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
font_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
font_body = Font(name=FONT_NAME, size=10, color="334155")
font_body_bold = Font(name=FONT_NAME, size=10, bold=True, color="0F172A")
font_formula = Font(name=FONT_NAME, size=10, bold=True, color="1E3A8A")

# Fills
fill_primary_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate 800
fill_accent_bar = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")      # Deep Teal 700
fill_row_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")        # Slate 50
fill_row_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")         # White
fill_formula_col = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")     # Light Blue 50

# Alert Fills & Fonts
fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")            # Red 100
font_red = Font(name=FONT_NAME, size=10, bold=True, color="991B1B")

fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")          # Green 100
font_green = Font(name=FONT_NAME, size=10, bold=True, color="15803D")

fill_orange = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")         # Orange 100
font_orange = Font(name=FONT_NAME, size=10, bold=True, color="C2410C")

# Borders
border_thin = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)
border_header = Border(
    left=Side(style='thin', color='475569'),
    right=Side(style='thin', color='475569'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='medium', color='1E293B')
)
border_total = Border(
    top=Side(style='thin', color='1E293B'),
    bottom=Side(style='double', color='1E293B')
)

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_sheet_header(ws, title, subtitle):
    """Writes a standard top title block to the sheet."""
    ws.views.sheetView[0].showGridLines = True
    
    # Title row
    ws.cell(row=1, column=1, value=title).font = font_title
    ws.row_dimensions[1].height = 28
    
    # Subtitle row
    ws.cell(row=2, column=1, value=subtitle).font = font_subtitle
    ws.row_dimensions[2].height = 18
    
    # Empty space row
    ws.row_dimensions[3].height = 12

def apply_alternating_rows(ws, start_row, end_row, max_col):
    """Applies clean slate-50 alternating row fills and thin borders."""
    for r in range(start_row, end_row + 1):
        is_even = r % 2 == 0
        row_fill = fill_row_even if is_even else fill_row_odd
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            # Preserve cell values and formulas, but set font/fill if not already stylized
            cell.border = border_thin
            if cell.fill.fill_type is None:
                cell.fill = row_fill
            if cell.font.name is None or cell.font.name != FONT_NAME:
                cell.font = font_body

def autofit_column_widths(ws, start_row=4, padding=4):
    """Autofits column widths dynamically with padding and safety bounds."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[start_row:]:
            if cell.value:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    max_len = max(max_len, 12)
                else:
                    lines = val_str.split('\n')
                    for line in lines:
                        max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = max(max_len + padding, 10)

# --- SLIDE DECK (PPT IN A SHEET) HELPER FUNCTIONS ---
def draw_slide_frame(ws, start_row, end_row, title, slide_num):
    """Creates a distinct slide outer frame with dark slate headers and light footer."""
    medium_side = Side(style='medium', color='1E293B')
    thin_side = Side(style='thin', color='CBD5E1')
    
    slide_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    # Fill slide workspace (Columns B to I)
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = slide_fill
            
            left = medium_side if c == 2 else None
            right = medium_side if c == 9 else None
            top = medium_side if r == start_row else None
            bottom = medium_side if r == end_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    # Title Bar (2 rows height merged)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row+1, end_column=9)
    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row+1].height = 20
    
    title_cell = ws.cell(row=start_row, column=2, value=title)
    title_cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = align_center
    
    # Border fixes for merged title cells
    for c in range(2, 10):
        cell_t = ws.cell(row=start_row, column=c)
        cell_t.fill = header_fill
        cell_t.border = Border(left=medium_side if c == 2 else None, 
                               right=medium_side if c == 9 else None, 
                               top=medium_side, 
                               bottom=thin_side)
        
        cell_b = ws.cell(row=start_row+1, column=c)
        cell_b.fill = header_fill
        cell_b.border = Border(left=medium_side if c == 2 else None, 
                               right=medium_side if c == 9 else None, 
                               top=thin_side, 
                               bottom=medium_side)
        
    # Slide Footer
    footer_row = end_row
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=9)
    ws.row_dimensions[footer_row].height = 20
    
    footer_cell = ws.cell(row=footer_row, column=2, value=f"SLIDE {slide_num} OF 5 | CHAPTER 2 DRIP SCALING STRATEGY PITCH")
    footer_cell.font = Font(name=FONT_NAME, size=8, italic=True, color="94A3B8")
    footer_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    for c in range(2, 10):
        cell = ws.cell(row=footer_row, column=c)
        cell.border = Border(left=medium_side if c == 2 else None, 
                             right=medium_side if c == 9 else None, 
                             top=thin_side, 
                             bottom=medium_side)

def draw_card(ws, start_row, end_row, start_col, end_col, title, text, header_fill):
    """Draws a themed card inside a slide frame with a colored top header."""
    thin_side = Side(style='thin', color='CBD5E1')
    
    # Card Header
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    hdr_cell = ws.cell(row=start_row, column=start_col, value=title)
    hdr_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="FFFFFF")
    hdr_cell.fill = header_fill
    hdr_cell.alignment = align_center
    
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = header_fill
        cell.border = Border(left=thin_side if c == start_col else None,
                             right=thin_side if c == end_col else None,
                             top=thin_side,
                             bottom=thin_side)
        
    # Card Body
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=end_row, end_column=end_col)
    body_cell = ws.cell(row=start_row+1, column=start_col, value=text)
    body_cell.font = Font(name=FONT_NAME, size=8.5, color="334155")
    body_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    body_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for r in range(start_row+1, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            left = thin_side if c == start_col else None
            right = thin_side if c == end_col else None
            top = thin_side if r == start_row+1 else None
            bottom = thin_side if r == end_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def create_presentation_sheet(ws):
    """Draws a Google Sheet-friendly slide deck template inside the sheet."""
    ws.views.sheetView[0].showGridLines = False
    
    # Setup slide widths B to I (Spacer A)
    ws.column_dimensions['A'].width = 3
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 18
        
    # --- SLIDE 1: COVER SLIDE (Rows 2 to 18) ---
    draw_slide_frame(ws, start_row=2, end_row=18, title="CHAPTER 2 DRIP: BOARD-LEVEL SCALING STRATEGY", slide_num=1)
    
    ws.merge_cells("B5:I5")
    sub_cell = ws.cell(row=5, column=2, value="An Algorithmic Growth Roadmap & Meta Ads Library Audit to Unlock ₹100 Cr+ Run Rate")
    sub_cell.font = Font(name=FONT_NAME, size=10.5, italic=True, color="475569")
    sub_cell.alignment = align_center
    
    goals_text = (
        "• Pivot acquisition from brand PR & founder dependency to data-driven performance ads.\n"
        "• Scale blended ad ROAS from 1.8x to 3.2x+ across Meta and Google.\n"
        "• Increase Average Order Value (AOV) to ₹3,200+ via bundle offers and sets."
    )
    draw_card(ws, start_row=8, end_row=12, start_col=2, end_col=5, title="Strategic Growth Goals", text=goals_text, header_fill=fill_accent_bar)
    
    levers_text = (
        "• Deploy Meta Advantage+ Shopping Campaigns (ASC+) for auto-targeting.\n"
        "• Integrate strict 'Drop Scarcity' & FOMO copywriting in ad copy.\n"
        "• Creative Factory system: Launch 10-15 new test variations weekly.\n"
        "• Google Search brand defense + generic PMax keywords setup."
    )
    draw_card(ws, start_row=8, end_row=12, start_col=6, end_col=9, title="Primary Operational Levers", text=levers_text, header_fill=fill_accent_bar)
    
    ws.merge_cells("B14:I15")
    ctx_text = (
        "Context & Vision:\n"
        "Co-founded in 2024 by Rhea & Showik Chakraborty. Currently valued at ₹40 Cr with a flagship experience store in Bandra, Mumbai. Built on the core values of reinvention, emotional storytelling, and premium unisex streetwear. Scaling now requires decoupling customer acquisition from co-founder celebrity buzz."
    )
    ctx_cell = ws.cell(row=14, column=2, value=ctx_text)
    ctx_cell.font = Font(name=FONT_NAME, size=9, color="475569")
    ctx_cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    ctx_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Fill background colors of context cell
    for r in range(14, 16):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            
    ws.row_dimensions[19].height = 15
    
    # --- SLIDE 2: UNIT ECONOMICS (Rows 20 to 34) ---
    draw_slide_frame(ws, start_row=20, end_row=34, title="SLIDE 2: CURRENT Baseline VS. TARGET ECONOMICS", slide_num=2)
    curr_econ_text = (
        "• Monthly Media Spend:  ₹90,00,000\n"
        "• Blended Ad ROAS:  1.8x\n"
        "• Gross Sales Revenue:  ₹1,62,00,000\n"
        "• Average Order Value:  ₹2,200\n"
        "• Blended COGS & Fulfillment:  52.8%\n"
        "• Blended CAC:  ₹1,222\n"
        "• CM1 Margin:  -₹13,50,000 (-8.3%)\n"
        "Observation: Low AOV + low ROAS restricts scale."
    )
    draw_card(ws, start_row=23, end_row=31, start_col=2, end_col=4, title="Current Baseline (Est)", text=curr_econ_text, header_fill=PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid"))
    
    targ_econ_text = (
        "• Monthly Media Spend:  ₹1,35,00,000\n"
        "• Blended Ad ROAS:  3.2x\n"
        "• Gross Sales Revenue:  ₹4,32,00,000\n"
        "• Average Order Value:  ₹3,200\n"
        "• Blended COGS & Fulfillment:  43.8%\n"
        "• Blended CAC:  ₹1,000\n"
        "• CM1 Margin:  +₹1,08,00,000 (+25.0%)\n"
        "Observation: Raising AOV to ₹3,200 unlocks media scale."
    )
    draw_card(ws, start_row=23, end_row=31, start_col=5, end_col=7, title="Target Scale Economics", text=targ_econ_text, header_fill=PatternFill(start_color="15803D", end_color="15803D", fill_type="solid"))
    
    insight_text = (
        "Why AOV is our bottleneck:\n"
        "At ₹2,200 AOV, the maximum CAC we can afford to hit a 3.0x ROAS is ₹733.\n\n"
        "By bundling products (Tee + Cargo) to raise AOV to ₹3,200, we raise our CAC tolerance to ₹1,000. This allows us to bid higher, outbid competitors in the Meta auction, and scale spending profitably."
    )
    draw_card(ws, start_row=23, end_row=31, start_col=8, end_col=9, title="Marketer Growth Insight", text=insight_text, header_fill=PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"))
    
    ws.row_dimensions[35].height = 15
    
    # --- SLIDE 3: ADS AUDIT GAPS (Rows 36 to 50) ---
    draw_slide_frame(ws, start_row=36, end_row=50, title="SLIDE 3: META ADS AUDIT - PRIMARY PERFORMANCE GAPS", slide_num=3)
    
    gap1_text = (
        "• Observation: Cinematic video ads of Rhea Chakraborty fatigue fast.\n"
        "• Impact: High cold acquisition costs.\n"
        "• Fix: Pivot to product-first UGC styling checks by micro-creators who look like the target buyer."
    )
    draw_card(ws, start_row=39, end_row=47, start_col=2, end_col=3, title="Gap 1: Celeb Dependency", text=gap1_text, header_fill=fill_primary_header)
    
    gap2_text = (
        "• Observation: Only a few active high-production video ads in the library.\n"
        "• Impact: Meta algorithm cannot optimize cohorts.\n"
        "• Fix: Establish a 'Creative Factory' launching 10-15 raw UGC variations weekly."
    )
    draw_card(ws, start_row=39, end_row=47, start_col=4, end_col=5, title="Gap 2: Low Asset Count", text=gap2_text, header_fill=fill_primary_header)
    
    gap3_text = (
        "• Observation: Ads focus on lifestyle; ignore fit, fabric weight, or GSM.\n"
        "• Impact: Low page-to-cart conversions.\n"
        "• Fix: Use text overlays detailing specs (e.g. 240 GSM, puff prints)."
    )
    draw_card(ws, start_row=39, end_row=47, start_col=6, end_col=7, title="Gap 3: Vibe-Only Focus", text=gap3_text, header_fill=fill_primary_header)
    
    gap4_text = (
        "• Concept: Streetwear buyers are driven by scarcity & FOMO.\n"
        "• Action: Run ad copy around drop availability: 'Drop 02 - Sold out means gone.'\n"
        "• Retargeting: Run countdown timer ads on Instagram Stories."
    )
    draw_card(ws, start_row=39, end_row=47, start_col=8, end_col=9, title="Quick Win: Drop Urgency", text=gap4_text, header_fill=PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid"))
    
    ws.row_dimensions[51].height = 15
    
    # --- SLIDE 4: COMPETITOR MATRIX (Rows 52 to 64) ---
    draw_slide_frame(ws, start_row=52, end_row=64, title="SLIDE 4: STREETWEAR COMPETITOR INTELLIGENCE MATRIX", slide_num=4)
    
    comp1_text = (
        "• Jaywalking (Premium):\n"
        "  - Tactic: High brand equity, model fit checks, raw street photos, zero discounts.\n"
        "  - Action for C2: Target metropolitan high-income areas with premium styling guides.\n"
        "• Bluorng (Premium):\n"
        "  - Tactic: Strict drop model (never restocked), puff print fabric texture close-ups.\n"
        "  - Action for C2: Design ad copy around scarcity: 'Drop 02 - Sold out means gone.'"
    )
    draw_card(ws, start_row=55, end_row=61, start_col=2, end_col=5, title="Premium Competitors: Jaywalking & Bluorng", text=comp1_text, header_fill=fill_accent_bar)
    
    comp2_text = (
        "• Bonkers Corner (Affordable):\n"
        "  - Tactic: Gen Z UGC, college try-on hauls, student style checks. High discount loops.\n"
        "  - Action for C2: Adopt try-on haul formats but retain C2's premium aesthetic.\n"
        "• Snitch (Fast Fashion):\n"
        "  - Tactic: AI-driven stock, 2-week shelf cycles, Advantage+ Shopping (ASC+) volume ads.\n"
        "  - Action for C2: Set up automated Meta ASC campaigns."
    )
    draw_card(ws, start_row=55, end_row=61, start_col=6, end_col=9, title="Volume Competitors: Bonkers Corner & Snitch", text=comp2_text, header_fill=fill_accent_bar)
    
    ws.row_dimensions[65].height = 15
    
    # --- SLIDE 5: MEDIA FUNNEL (Rows 66 to 77) ---
    draw_slide_frame(ws, start_row=66, end_row=77, title="SLIDE 5: THE 4-STAGE SCALING MEDIA FUNNEL (ROAS 3+)", slide_num=5)
    
    fnl1_text = (
        "• Target: Broad Targeting & Streetwear interest stack.\n"
        "• Creative: Transformation reels, fabric spec guides, unboxing hooks.\n"
        "• KPIs: CTR > 2.0%, Hook > 35%."
    )
    draw_card(ws, start_row=69, end_row=74, start_col=2, end_col=3, title="TOF: Cold Acquisition (70%)", text=fnl1_text, header_fill=fill_primary_header)
    
    fnl2_text = (
        "• Target: Page Engagers & 50%+ Video Viewers (exclude buyers).\n"
        "• Creative: Haul reviews, styling guides ('5 Ways to Style').\n"
        "• KPIs: CTR > 3.0%, ATC Rate > 8%."
    )
    draw_card(ws, start_row=69, end_row=74, start_col=4, end_col=5, title="MOFU: Consideration (15%)", text=fnl2_text, header_fill=fill_primary_header)
    
    fnl3_text = (
        "• Target: Cart Abandoners & Checkouts (30 days).\n"
        "• Creative: Catalog DPAs + checkout bundle offers.\n"
        "• KPIs: ROAS > 4.5x, Conv. > 3.5%."
    )
    draw_card(ws, start_row=69, end_row=74, start_col=6, end_col=7, title="BOF: Direct Conversion (10%)", text=fnl3_text, header_fill=fill_primary_header)
    
    fnl4_text = (
        "• Target: Past Customers (365 days).\n"
        "• Creative: VIP early access (48-hour drop headstart).\n"
        "• KPIs: Repeat Purchase > 25%."
    )
    draw_card(ws, start_row=69, end_row=74, start_col=8, end_col=9, title="Retention & LTV (5%)", text=fnl4_text, header_fill=PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"))

def create_audit_sheet(ws):
    write_sheet_header(ws, "Chapter 2 Drip - Meta Ads Library Audit & Gaps Analysis", 
                       "Conducted by Antigravity D2C Growth Command Center | Target: Identify Gaps to Unlock Scale")
    
    headers = [
        "S.No.", "Key Audited Aspect", "Current Observation (What they are doing)", 
        "Identified Gap (What is missing)", "Scaling Impact", "D2C Marketer Strategic Recommendation", "Priority"
    ]
    
    ws.row_dimensions[4].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary_header
        cell.alignment = align_header
        cell.border = border_header
        
    audit_data = [
        (
            "Founder-Centric Creatives",
            "Heavy reliance on co-founder Rhea Chakraborty's personal brand, aesthetics, and direct features in high-budget short films.",
            "Ads suffer from rapid creative fatigue. Personal founder narratives have high PR value but fail to scale cold programmatic audiences who don't follow celebrity gossip.",
            "High",
            "De-risk by introducing product-first and consumer-first creatives. Implement styling tutorials, street-style hauls, and fit-checks by real Gen Z creators who look like the target buyer.",
            "P0"
        ),
        (
            "Low Creative Velocity",
            "Very low volume of active ad assets in the library. Pushing a few highly polished, cinematic hero assets (like Tinder Holi collab or 'Denims & Delusions').",
            "Meta's auction algorithm requires a high volume of raw creative inputs to test different hooks and match the right ad to the right cohort. Low asset counts lead to high CPA.",
            "High",
            "Establish a 'Creative Factory' system. Partner with 40-50 micro-influencers monthly to generate raw visual hooks. Target releasing 10-15 new ad variations weekly in various hooks/formats.",
            "P0"
        ),
        (
            "Lack of Product Feature Focus",
            "Creatives focus purely on 'vibe', emotional messaging (e.g. resilience, LGBTQ+ acceptance) and moody lifestyle scenes.",
            "At a premium AOV of ₹2,500+, the consumer needs to understand the product value: fabric weight (GSM), comfort, wash durability, and detailing (e.g., modular pockets).",
            "High",
            "Launch structured product-demystifier ads. Use text overlays and voiceovers detailing product specs: '240 GSM Heavyweight French Terry Cotton', 'Acid-Wash Distressed Denim', fit charts.",
            "P0"
        ),
        (
            "No Dynamic Retargeting (DPA)",
            "Active ads are primarily single-video reels pointing to the homepage or collection page. No active Advantage+ Catalog Ads running.",
            "High cart-abandonment rate. Streetwear is an impulse purchase, and warm traffic is falling out of the funnel without personalized product reminders.",
            "High",
            "Deploy Dynamic Product Ads (DPAs) and Advantage+ Catalog campaigns. Retarget add-to-carts with specific carousel cards displaying the exact products viewed, combined with a 10% first-purchase coupon.",
            "P1"
        ),
        (
            "Absence of Scarcity / Drop Mechanics",
            "Drops are highlighted on social feed but ads run indefinitely without urgency banners.",
            "Fails to build the psychological FOMO (Fear of Missing Out) that drives high-ROAS streetwear brands (like Bluorng or Supreme).",
            "High",
            "Inject Drop Mechanics into ad copy and creatives: 'Drop 02: Only 500 units available. No Restocks.' or '92% Sold Out - Final Restock warning.' Run countdown timers on stories.",
            "P0"
        ),
        (
            "High-Intent Search Leakage",
            "Almost 100% of budget spent on Meta/Instagram. Minimal Google Search or Performance Max coverage.",
            "Competitors (like Snitch and Bonkers) are bidding on Chapter 2 brand terms and high-intent generic keywords (e.g. 'premium cargos India', 'unisex streetwear Mumbai'), stealing high-intent buyers.",
            "Medium",
            "Set up a Google PMax campaign targeting high-intent keywords, and set up a Brand Search campaign to protect brand search real estate. Secure conversions from customers searching the brand directly.",
            "P1"
        ),
        (
            "Strictly High-Brow English Copy",
            "Creative copy, voiceovers, and captions are exclusively in written English with sophisticated language.",
            "Misses massive regional streetwear pockets in Punjab, Pune, Northeast, and tier-2 colleges where streetwear is booming and relatable Hinglish/regional copy converts better.",
            "Medium",
            "Run parallel local-concept creatives. Use Hinglish voiceovers ('Drip level max, price range sorted') and local culture hooks. Test regional localization for Delhi NCR vs Mumbai street culture.",
            "P2"
        ),
        (
            "Post-Click Friction (CRO)",
            "Ad clicks redirect to standard, unoptimized Shopify product pages. Size guides are standard text charts; no interactive bundle builders.",
            "High traffic drop-off at checkout. Streetwear buyers buy sets (hoodie + jogger) but have to navigate to separate pages, lowering the AOV and ROAS.",
            "High",
            "Build high-converting product pages with video overlays. Implement 'Shop the Look' bundle discounts directly on product pages (e.g. 'Add Joggers for 20% off'). This raises AOV, allowing us to bid higher on Meta.",
            "P1"
        )
    ]
    
    current_row = 5
    for idx, row_data in enumerate(audit_data, 1):
        ws.row_dimensions[current_row].height = 55
        
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        
        cell_aspect = ws.cell(row=current_row, column=2, value=row_data[0])
        cell_aspect.font = font_body_bold
        cell_aspect.alignment = align_left
        
        ws.cell(row=current_row, column=3, value=row_data[1]).alignment = align_left
        ws.cell(row=current_row, column=4, value=row_data[2]).alignment = align_left
        
        cell_impact = ws.cell(row=current_row, column=5, value=row_data[3])
        cell_impact.alignment = align_center
        if row_data[3] == "High":
            cell_impact.fill = fill_red
            cell_impact.font = font_red
        else:
            cell_impact.fill = fill_orange
            cell_impact.font = font_orange
            
        ws.cell(row=current_row, column=6, value=row_data[4]).alignment = align_left
        
        cell_priority = ws.cell(row=current_row, column=7, value=row_data[5])
        cell_priority.alignment = align_center
        if row_data[5] == "P0":
            cell_priority.fill = fill_red
            cell_priority.font = font_red
        elif row_data[5] == "P1":
            cell_priority.fill = fill_orange
            cell_priority.font = font_orange
        else:
            cell_priority.fill = fill_green
            cell_priority.font = font_green
            
        current_row += 1
        
    apply_alternating_rows(ws, 5, current_row - 1, 7)
    autofit_column_widths(ws, start_row=4)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 12
    ws.freeze_panes = "C5"

def create_competitor_sheet(ws):
    write_sheet_header(ws, "Streetwear Competitor Intelligence Matrix", 
                       "Analyzing the Ads Strategy of Key D2C Streetwear Brands in the Indian Market")
    
    headers = [
        "S.No.", "Competitor", "Price Segment", "Core Strengths & Positioning", 
        "Ads Creative Strategy", "Media Buying Channels", "Key Actionable Takeaway for Chapter 2"
    ]
    
    ws.row_dimensions[4].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_accent_bar
        cell.alignment = align_header
        cell.border = border_header
        
    comp_data = [
        (
            "Jaywalking",
            "Ultra-Premium\n(Tees: ₹3k-6k, Cargo/Jackets: ₹8k+)",
            "Pioneered Indian luxury D2C streetwear. Cult community. Handcrafted avant-garde shapes & high-art branding.",
            "Raw model lookbooks, urban background styling, zero-discount, focus on silhouette shape and structural drops.",
            "Meta Ads (Awareness & Conversions), organic Instagram seeding on rappers and designers.",
            "Target highly metropolitan, high-income zip codes with raw styling guides. Don't run standard conversion discounts; focus on premium silhouette differentiation."
        ),
        (
            "Bluorng",
            "Premium Streetwear\n(Tees: ₹3.5k-5k, Hoodies: ₹5k-7k)",
            "Strict Drop Model (never restocked). High lifestyle aspiration, intense celebrity/creator gifting seeding.",
            "High-contrast lifestyle videos, extreme closeups of puff printing, high GSM puff fabric texture, and model fit transitions.",
            "Meta Ads (high creative variety), Google PMax, SEO (heavy investment to build organic brand trust).",
            "Adopt the drop model in performance ads. Design ad copy around scarcity: 'Drop 02: Only 500 units. Sold out means gone.' Showcase fabric details close-up to justify ₹3k+ prices."
        ),
        (
            "Bonkers Corner",
            "Affordable / Mass\n(Tees: ₹799-1,499, Pants: ₹1,500+)",
            "Pop-culture licensing (Marvel, Disney, Anime), hyper-rapid catalog expansion, very low barrier to entry.",
            "High-volume Gen Z UGC, college try-on hauls, reaction videos, and trend-jacking. Focuses on humor, comfort, and discounts.",
            "Meta Ads, Snapchat Ads (heavy budget targeting teens), TikTok/Reels organic viral reach.",
            "Leverage their UGC formats (try-on hauls, student fit checks) but maintain Chapter 2's premium, edgy aesthetic. Do not run mass discount ads that cheapen the celebrity association."
        ),
        (
            "Snitch",
            "Fast Fashion Streetwear\n(Tees: ₹999-1,999, Shirts: ₹1,499+)",
            "AI-driven supply chain, 2-week shelf cycle, highly systemized performance marketing, Shark Tank credibility.",
            "Relatable lifestyle reels, fast-paced transitions, humor/meme hooks, Advantage+ Shopping (ASC) volume catalog ads.",
            "Meta Ads (heavy Advantage+ focus), Google Ads (PMax & high search bid share), YouTube pre-rolls.",
            "Build an Advantage+ Shopping (ASC+) campaign structure. Shift budget from cold interests to broad targeting. Launch daily catalog ad variations. Standardize nomenclature."
        )
    ]
    
    current_row = 5
    for idx, row_data in enumerate(comp_data, 1):
        ws.row_dimensions[current_row].height = 75
        
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        
        cell_name = ws.cell(row=current_row, column=2, value=row_data[0])
        cell_name.font = font_body_bold
        cell_name.alignment = align_center
        
        ws.cell(row=current_row, column=3, value=row_data[1]).alignment = align_center
        ws.cell(row=current_row, column=4, value=row_data[2]).alignment = align_left
        ws.cell(row=current_row, column=5, value=row_data[3]).alignment = align_left
        ws.cell(row=current_row, column=6, value=row_data[4]).alignment = align_left
        ws.cell(row=current_row, column=7, value=row_data[5]).alignment = align_left
        
        current_row += 1
        
    apply_alternating_rows(ws, 5, current_row - 1, 7)
    autofit_column_widths(ws, start_row=4)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 45
    ws.freeze_panes = "C5"

def create_strategy_sheet(ws):
    write_sheet_header(ws, "Chapter 2 Drip - 3.0+ ROAS Strategic Blueprint", 
                       "Full-Funnel Media Buying Blueprint & Creative Directive for Scaling Streetwear to ₹10 Cr+ Run Rate")
    
    headers = [
        "Funnel Stage", "Target Audience Group", "Budget Allocation", 
        "Creative Strategy & Assets", "Core Key Performance Indicators (KPIs)", "Step-by-Step Action Plan"
    ]
    
    ws.row_dimensions[4].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary_header
        cell.alignment = align_header
        cell.border = border_header
        
    strategy_data = [
        (
            "Top of Funnel (TOF)\n- Cold Acquisition",
            "Broad Targeting (No interests, let Meta's pixel optimize).\nInterest Stacks: Streetwear, Sneakers, Indian Hip Hop, College Pockets (18-30, Metros & Tier-1).",
            "70% of Total Budget",
            "• Hook-Heavy Reels: Quick transitions showing streetwear transformations (e.g., 'From raw look to drop style in 5 seconds').\n• Quality Demystifiers: Show close-ups of premium fabric texture and stitch detailing.\n• Storytelling Reels: Founder Rhea sharing the brand's 'resilience/second chapter' message.",
            "• Outbound CTR: > 2.0%\n• Hook Rate (3-sec play / Imp): > 35%\n• Hold Rate (15-sec play / 3-sec play): > 20%\n• Blended CPC: < ₹25",
            "1. Launch 1 Meta Advantage+ Shopping Campaign (ASC).\n2. Partner with 30 micro-creators monthly to feed the ASC creative machine.\n3. Launch 3 new creative concepts weekly to test different hooks."
        ),
        (
            "Middle of Funnel (MOFU)\n- Consideration",
            "Engagers Retargeting: 180-day Instagram & Facebook page engagers, 50%+ video viewers of previous campaigns.",
            "15% of Total Budget",
            "• Social Proof UGC: Unboxing videos and raw streetwear haul reviews showing real customer fits.\n• Styling Guides: '5 Ways to Style Our Acid Wash Denim Jacket.'\n• Store Experience: Coffee nook 'Chapter 2 Brew' highlights + store walk-throughs in Bandra to build local trust.",
            "• CTR: > 3.0%\n• Landing Page View Rate: > 75%\n• Add to Cart (ATC) Rate: > 8%\n• Cost per ATC: < ₹150",
            "1. Run a custom audience retargeting ad set.\n2. Exclude past purchasers of the last 30 days.\n3. Dynamic optimization: use interactive carousel ads with customer reviews overlayed."
        ),
        (
            "Bottom of Funnel (BOF)\n- Direct Conversion",
            "Cart Abandoners & Checkout Initiators: 30-day website visitors who added to cart but did not purchase.",
            "10% of Total Budget",
            "• Dynamic Product Ads (DPA): Carousels displaying exact product items left in cart.\n• Bundle Offer Ads: 'Complete the Drip: Buy Tee + Cargo, save ₹500.'\n• Trust Alerts: Clear banners highlighting Free Shipping, Easy 7-Day Exchange, Cash on Delivery (COD).",
            "• ROAS: > 4.5x\n• Purchase Conversion Rate: > 3.5%\n• Cost Per Purchase (CPP): < ₹800",
            "1. Deploy a catalog sales campaign mapped directly to the shopify product feed.\n2. Set up automated overlays showing positive review stars on product catalog images.\n3. Offer an automatic cart-recovery discount at checkout."
        ),
        (
            "Retention & LTV\n- Brand Advocates",
            "Loyalty Retargeting: Past buyers who purchased in the last 365 days.",
            "5% of Total Budget",
            "• VIP Early Access: Ads highlighting collection drops live 48 hours early for VIPs only.\n• Cross-sell Packs: Retargeting t-shirt buyers with utility cargo pants and premium accessories (caps, socks).\n• Community Building: Invitations to exclusive meetups/events at the Bandra flagship store.",
            "• Repeat Purchase Rate: > 25%\n• LTV Growth: +30% MoM\n• CAC payback period: Instant",
            "1. Create a Custom Audience of past purchasers.\n2. Exclude them from acquisition campaigns to save budget.\n3. Run early-bird preview campaigns for new drops 2 days before public release."
        )
    ]
    
    current_row = 5
    for idx, row_data in enumerate(strategy_data, 1):
        ws.row_dimensions[current_row].height = 110
        
        cell_stage = ws.cell(row=current_row, column=1, value=row_data[0])
        cell_stage.font = font_body_bold
        cell_stage.alignment = align_center
        
        ws.cell(row=current_row, column=2, value=row_data[1]).alignment = align_left
        
        cell_budget = ws.cell(row=current_row, column=3, value=row_data[2])
        cell_budget.alignment = align_center
        cell_budget.font = font_body_bold
        cell_budget.fill = fill_formula_col
        
        ws.cell(row=current_row, column=4, value=row_data[3]).alignment = align_left
        ws.cell(row=current_row, column=5, value=row_data[4]).alignment = align_left
        ws.cell(row=current_row, column=6, value=row_data[5]).alignment = align_left
        
        current_row += 1
        
    apply_alternating_rows(ws, 5, current_row - 1, 6)
    autofit_column_widths(ws, start_row=4)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 45
    ws.freeze_panes = "B5"

def create_pl_simulator_sheet(ws):
    write_sheet_header(ws, "Chapter 2 Drip - Growth & Unit Economics Simulator", 
                       "Interactive Simulator: Modify input cells in light orange/white to model path to 3.0+ ROAS")
    
    ws.row_dimensions[4].height = 24
    ws.merge_cells("A4:F4")
    cell_sec1 = ws.cell(row=4, column=1, value="1. Media Performance Inputs (Directly Edit White/Orange Cells)")
    cell_sec1.font = font_section_header
    cell_sec1.fill = fill_accent_bar
    cell_sec1.alignment = align_left
    
    ws.row_dimensions[5].height = 24
    subheaders = ["Variable Name", "Current State (Est)", "Target State (Scale)", "Formula/Logic Reference", "Strategic Growth Lever", "Unit / Format"]
    for col_idx, sh in enumerate(subheaders, 1):
        cell = ws.cell(row=5, column=col_idx, value=sh)
        cell.font = font_header
        cell.fill = fill_primary_header
        cell.alignment = align_header
        cell.border = border_header
        
    vars_data = [
        ("Average Monthly Media Spend", 9000000, 13500000, "Direct Input", "Ad scaling budget allocation based on funding & cash flow", "INR"),
        ("Average Order Value (AOV)", 2200, 3200, "Direct Input", "Increase via bundle offers, upsells at checkout, pricing updates", "INR"),
        ("Blended COGS & Fulfillment %", 0.5277, 0.4375, "Direct Input", "Decrease via bulk sourcing, prepaid incentives, regional warehouses", "PERCENT"),
        ("Ad Spend ROAS", 1.8, 3.2, "Direct Input", "Increase via high-angle creative velocity, ASC implementation", "MULTIPLIER"),
        ("Gross Sales Revenue", "=B6*B9", "=C6*C9", "Media Spend * ROAS", "Dynamic outcome representing monthly D2C website sales", "INR_FORMULA"),
        ("Total COGS Amount", "=B10*B8", "=C10*C8", "Sales Revenue * COGS %", "Dynamic manufacturing cost of sold inventory", "INR_FORMULA"),
        ("Gross Margin Value", "=B10-B11", "=C10-C11", "Sales Revenue - Total COGS", "Dynamic profit before advertising costs", "INR_FORMULA"),
        ("Blended Target CAC", "=B7/B9", "=C7/C9", "AOV / ROAS", "Calculated cost to acquire 1 new purchasing customer", "INR_FORMULA"),
        ("Total Customers Acquired", "=B6/B13", "=C6/C13", "Media Spend / CAC", "Total transactional volume generated per month", "INT_FORMULA"),
        ("Contribution Margin 1 (CM1)", "=B12-B6", "=C12-C6", "Gross Margin - Media Spend", "D2C Profitability metric after media buying costs", "INR_FORMULA"),
        ("Contribution Margin 1 % (CM1 %)", "=B15/B10", "=C15/C10", "CM1 / Sales Revenue", "Relative efficiency of the business scaling", "PERCENT_FORMULA")
    ]
    
    current_row = 6
    for v in vars_data:
        ws.row_dimensions[current_row].height = 28
        
        cell_name = ws.cell(row=current_row, column=1, value=v[0])
        cell_name.font = font_body_bold
        cell_name.alignment = align_left
        
        cell_curr = ws.cell(row=current_row, column=2, value=v[1])
        cell_curr.alignment = align_right
        
        cell_targ = ws.cell(row=current_row, column=3, value=v[2])
        cell_targ.alignment = align_right
        
        ws.cell(row=current_row, column=4, value=v[3]).alignment = align_left
        ws.cell(row=current_row, column=5, value=v[4]).alignment = align_left
        
        fmt = v[5]
        if fmt == "INR":
            cell_curr.number_format = '₹#,##0'
            cell_curr.font = font_body
            cell_targ.number_format = '₹#,##0'
            cell_targ.font = font_body
        elif fmt == "PERCENT":
            cell_curr.number_format = '0.0%'
            cell_curr.font = font_body
            cell_targ.number_format = '0.0%'
            cell_targ.font = font_body
        elif fmt == "MULTIPLIER":
            cell_curr.number_format = '0.0"x"'
            cell_curr.font = font_body
            cell_targ.number_format = '0.0"x"'
            cell_targ.font = font_body
        elif fmt == "INR_FORMULA":
            cell_curr.number_format = '₹#,##0'
            cell_curr.font = font_formula
            cell_curr.fill = fill_formula_col
            cell_targ.number_format = '₹#,##0'
            cell_targ.font = font_formula
            cell_targ.fill = fill_formula_col
        elif fmt == "PERCENT_FORMULA":
            cell_curr.number_format = '0.0%'
            cell_curr.font = font_formula
            cell_curr.fill = fill_formula_col
            cell_targ.number_format = '0.0%'
            cell_targ.font = font_formula
            cell_targ.fill = fill_formula_col
        elif fmt == "INT_FORMULA":
            cell_curr.number_format = '#,##0'
            cell_curr.font = font_formula
            cell_curr.fill = fill_formula_col
            cell_targ.number_format = '#,##0'
            cell_targ.font = font_formula
            cell_targ.fill = fill_formula_col
            
        current_row += 1
        
    apply_alternating_rows(ws, 6, current_row - 1, 6)
    
    for col in range(1, 7):
        ws.cell(row=15, column=col).border = Border(top=Side(style='thin', color='1E293B'))
        cell_cm_pct = ws.cell(row=16, column=col)
        cell_cm_pct.border = border_total
        
    ws.cell(row=16, column=2).fill = fill_red
    ws.cell(row=16, column=2).font = font_red
    ws.cell(row=16, column=3).fill = fill_green
    ws.cell(row=16, column=3).font = font_green
    
    autofit_column_widths(ws, start_row=4)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 15
    ws.freeze_panes = "D6"

def build_workbook():
    print("Initializing Workbook Generation...")
    wb = Workbook()
    
    # Sheet 0 (Default Active Sheet) -> Slide Presentation tab
    ws_pres = wb.active
    ws_pres.title = "0. Slide Presentation (PPT)"
    create_presentation_sheet(ws_pres)
    
    # Sheet 1
    ws_audit = wb.create_sheet(title="1. Ads Library Audit")
    create_audit_sheet(ws_audit)
    
    # Sheet 2
    ws_comp = wb.create_sheet(title="2. Competitor Benchmarking")
    create_competitor_sheet(ws_comp)
    
    # Sheet 3
    ws_strat = wb.create_sheet(title="3. 3+ ROAS Blueprint")
    create_strategy_sheet(ws_strat)
    
    # Sheet 4
    ws_pl = wb.create_sheet(title="4. P&L & Scale Simulator")
    create_pl_simulator_sheet(ws_pl)
    
    # Save Workbook
    output_path = os.path.join(os.getcwd(), FILE_NAME)
    wb.save(output_path)
    print(f"Excel Audit Workbook successfully compiled and saved to: {output_path}")

if __name__ == "__main__":
    build_workbook()
