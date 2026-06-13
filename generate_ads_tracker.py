import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
FILE_NAME = "Ads_Tracker_with_Nomenclature.xlsx"

# Colors
HEADER_FILL_COLOR = "1E293B"  # Slate 800
HEADER_TEXT_COLOR = "FFFFFF"  # White
ROW_FILL_EVEN = "F8FAFC"      # Slate 50
ROW_FILL_ODD = "FFFFFF"       # White
BORDER_COLOR = "E2E8F0"       # Slate 200

# Sheet 1: Ads Tracker
ADS_COLUMNS = [
    ("S.No.", 8),
    ("Date Added", 15),
    ("Platform", 15),
    ("Product Name", 25),
    ("Ad Name", 40),
    ("Google Drive Link (4:5 / 1:1)", 35),
    ("Google Drive Link (9:16)", 35),
    ("Live?", 10),
    ("Can take live?", 15),
    ("Raised By", 20),
    ("Persona", 20),
    ("INT / INF", 15),
    ("Ad Type", 20),
    ("Broad Narrative - P0", 25),
    ("Narrative - P1", 25),
    ("Language", 15),
    ("Person Full Name", 20),
    ("Ad Format", 15),
    ("Characters", 15),
    ("Additional information", 40)
]

ADS_VALIDATIONS = {
    "Platform": '"Meta,Google"',
    "Live?": '"Yes,No"',
    "Can take live?": '"Yes,No"',
    "INT / INF": '"Int,Inf"',
    "Ad Type": '"Static,Video,Reel,Carousel,Gif,Boomerang,Motion Graphic"',
    "Language": '"English,Hindi,Hinglish,Marathi,Tamil,Telugu,Kannada,Malayalam"',
    "Ad Format": '"1:1,4:5,9:16,Story,Reel"'
}

# Sheet 2: Campaign Nomenclature
CAMP_COLUMNS = [
    ("S.No.", 8),
    ("Objective", 20),
    ("Funnel Stage", 20),
    ("Product Category", 25),
    ("Geography", 20),
    ("Generated Campaign Name", 50)
]

CAMP_VALIDATIONS = {
    "Objective": '"Conversions,Traffic,Awareness,Engagement,Leads"',
    "Funnel Stage": '"Prospecting,Retargeting,Retention,ASC+,Advantage+"',
    "Product Category": '"Haircare,Skincare,Bodycare,Supplements,Kits"',
    "Geography": '"Pan India,Tier 1,Tier 2,Tier 3,Metro"'
}

# Sheet 3: Adset Nomenclature
ADSET_COLUMNS = [
    ("S.No.", 8),
    ("Audience Type", 20),
    ("Audience Details", 30),
    ("Placement", 20),
    ("Bidding Strategy", 20),
    ("Generated Adset Name", 50)
]

ADSET_VALIDATIONS = {
    "Audience Type": '"Broad,Lookalike,Custom,Interest,Advantage+"',
    "Placement": '"Auto,IG Reels,FB Feed,Stories,Search,Display"',
    "Bidding Strategy": '"Highest Volume,Cost Cap,ROAS Goal,Target CPA"'
}

def apply_styles(ws, columns, validations, formula_col_idx=None, formula_template=None):
    # Define styles
    header_font = Font(name="Inter", size=11, bold=True, color=HEADER_TEXT_COLOR)
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Inter", size=10)
    formula_font = Font(name="Inter", size=10, bold=True, color="0F172A")
    cell_align_left = Alignment(horizontal="left", vertical="center")
    cell_align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    even_fill = PatternFill(start_color=ROW_FILL_EVEN, end_color=ROW_FILL_EVEN, fill_type="solid")
    odd_fill = PatternFill(start_color=ROW_FILL_ODD, end_color=ROW_FILL_ODD, fill_type="solid")
    formula_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # slightly darker for formula
    
    # 1. Write Headers & Set Column Widths
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    
    # 2. Add Data Validations
    dv_dict = {}
    for col_name, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select a value from the list."
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select from list"
        dv.promptTitle = "Selection"
        ws.add_data_validation(dv)
        dv_dict[col_name] = dv

    # 3. Format Rows & Apply Validations / Formulas
    TOTAL_ROWS = 500
    for row_idx in range(2, TOTAL_ROWS + 2):
        ws.row_dimensions[row_idx].height = 25
        is_even = row_idx % 2 == 0
        current_fill = even_fill if is_even else odd_fill
        
        for col_idx, (col_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Apply styles
            if col_idx == formula_col_idx:
                cell.font = formula_font
                cell.fill = formula_fill
                # Inject formula e.g. =IF(COUNTA(B2:E2)>0, TEXTJOIN("_", TRUE, B2:E2), "")
                cell.value = formula_template.format(row=row_idx)
            else:
                cell.font = cell_font
                cell.fill = current_fill
            
            cell.border = thin_border
            
            # Center alignment for specific short columns
            if col_name in ["S.No.", "Date Added", "Platform", "Live?", "Can take live?", "INT / INF", "Language", "Ad Format"]:
                cell.alignment = cell_align_center
            else:
                cell.alignment = cell_align_left
                
            # Apply data validation if exists
            if col_name in dv_dict:
                col_letter = get_column_letter(col_idx)
                cell_coord = f"{col_letter}{row_idx}"
                dv_dict[col_name].add(cell_coord)
                
            # Auto-increment S.No.
            if col_name == "S.No.":
                cell.value = row_idx - 1

def create_dashboard():
    print(f"Generating Premium Ads & Nomenclature Tracker: {FILE_NAME}")
    
    wb = Workbook()
    
    # Sheet 1: Ads Tracker
    ws_ads = wb.active
    ws_ads.title = "Ads Tracker"
    apply_styles(ws_ads, ADS_COLUMNS, ADS_VALIDATIONS)
    
    # Sheet 2: Campaign Nomenclature
    ws_camp = wb.create_sheet("Campaign Nomenclature")
    # Formula to join B, C, D, E with underscores
    camp_formula = '=IF(COUNTA(B{row}:E{row})>0, TEXTJOIN("_", TRUE, B{row}:E{row}), "")'
    apply_styles(ws_camp, CAMP_COLUMNS, CAMP_VALIDATIONS, formula_col_idx=6, formula_template=camp_formula)
    
    # Sheet 3: Adset Nomenclature
    ws_adset = wb.create_sheet("Adset Nomenclature")
    # Formula to join B, C, D, E with underscores
    adset_formula = '=IF(COUNTA(B{row}:E{row})>0, TEXTJOIN("_", TRUE, B{row}:E{row}), "")'
    apply_styles(ws_adset, ADSET_COLUMNS, ADSET_VALIDATIONS, formula_col_idx=6, formula_template=adset_formula)
    
    # Save the Workbook
    output_path = os.path.join(os.getcwd(), FILE_NAME)
    wb.save(output_path)
    print(f"Success! Saved to {output_path}")

if __name__ == "__main__":
    create_dashboard()
