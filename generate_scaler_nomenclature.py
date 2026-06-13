import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FILE_NAME = "D2C_Scaler_Ads_Nomenclature.xlsx"

# Colors
HEADER_FILL_COLOR = "1E293B"  # Slate 800
HEADER_TEXT_COLOR = "FFFFFF"  # White
ROW_FILL_EVEN = "F8FAFC"      # Slate 50
ROW_FILL_ODD = "FFFFFF"       # White
BORDER_COLOR = "E2E8F0"       # Slate 200

def apply_styles(ws, columns, validations, formula_col_idx=None, formula_template=None):
    header_font = Font(name="Inter", size=11, bold=True, color=HEADER_TEXT_COLOR)
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Inter", size=10)
    formula_font = Font(name="Inter", size=10, bold=True, color="0F172A")
    cell_align_left = Alignment(horizontal="left", vertical="center")
    cell_align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    even_fill = PatternFill(start_color=ROW_FILL_EVEN, end_color=ROW_FILL_EVEN, fill_type="solid")
    odd_fill = PatternFill(start_color=ROW_FILL_ODD, end_color=ROW_FILL_ODD, fill_type="solid")
    formula_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    # 1. Write Headers & Set Column Widths
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    
    # 2. Add Data Validations
    dv_dict = {}
    for col_name, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select a value from the list."
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select from list"
        dv.promptTitle = "Selection"
        ws.add_data_validation(dv)
        dv_dict[col_name] = dv

    # 3. Format Rows & Apply Validations / Formulas
    TOTAL_ROWS = 1000
    for row_idx in range(2, TOTAL_ROWS + 2):
        ws.row_dimensions[row_idx].height = 25
        is_even = row_idx % 2 == 0
        current_fill = even_fill if is_even else odd_fill
        
        for col_idx, (col_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_idx == formula_col_idx:
                cell.font = formula_font
                cell.fill = formula_fill
                cell.value = formula_template.format(row=row_idx)
            else:
                cell.font = cell_font
                cell.fill = current_fill
            
            cell.border = thin_border
            
            if col_name == "S.No.":
                cell.alignment = cell_align_center
                cell.value = row_idx - 1
            else:
                cell.alignment = cell_align_left
                
            if col_name in dv_dict:
                col_letter = get_column_letter(col_idx)
                cell_coord = f"{col_letter}{row_idx}"
                dv_dict[col_name].add(cell_coord)

def create_dashboard():
    print(f"Generating D2C Scaler Nomenclature Tracker: {FILE_NAME}")
    wb = Workbook()
    
    # ---------------------------------------------------------
    # 1. META CAMPAIGN NOMENCLATURE
    # [Platform]_[Funnel Stage]_[Objective]_[Product/Category]_[Promo/Evergreen]_[Geo]
    # ---------------------------------------------------------
    ws_meta_camp = wb.active
    ws_meta_camp.title = "Meta Campaign"
    
    META_CAMP_COLS = [
        ("S.No.", 8),
        ("Platform", 15),
        ("Funnel Stage", 20),
        ("Objective", 20),
        ("Product/Category", 25),
        ("Promo/Evergreen", 20),
        ("Geo", 20),
        ("Generated Campaign Name", 70)
    ]
    META_CAMP_VALS = {
        "Platform": '"Meta"',
        "Funnel Stage": '"TOF,MOF,BOF,Retention,ASC+"',
        "Objective": '"Conversions,Traffic,Lead Gen,Catalog Sales,Engagement,Awareness"',
        "Product/Category": '"Haircare,Skincare,Bodycare,Supplements,Kits,All Products"',
        "Promo/Evergreen": '"Evergreen,BOGO,Flat50,Festive,Flash Sale"',
        "Geo": '"Pan India,Tier 1,Tier 2,Metro,North,South,US,Global"'
    }
    meta_camp_formula = '=IF(COUNTA(B{row}:G{row})>0, TEXTJOIN("_", TRUE, B{row}:G{row}), "")'
    apply_styles(ws_meta_camp, META_CAMP_COLS, META_CAMP_VALS, formula_col_idx=8, formula_template=meta_camp_formula)

    # ---------------------------------------------------------
    # 2. META ADSET NOMENCLATURE
    # [Audience Type]_[Audience Details]_[Demographics]_[Placement]
    # ---------------------------------------------------------
    ws_meta_adset = wb.create_sheet("Meta Adset")
    META_ADSET_COLS = [
        ("S.No.", 8),
        ("Audience Type", 20),
        ("Audience Details (e.g., Purchasers180D, 1-3%)", 40),
        ("Demographics", 20),
        ("Placement", 20),
        ("Generated Adset Name", 70)
    ]
    META_ADSET_VALS = {
        "Audience Type": '"Broad,Lookalike,Custom,Interest,Open"',
        "Demographics": '"18-35_F,25-44_M,18-65_All,18-24_F"',
        "Placement": '"Automatic,FB_Feed,IG_Reels,IG_Stories,Manual_FB_IG"'
    }
    meta_adset_formula = '=IF(COUNTA(B{row}:E{row})>0, TEXTJOIN("_", TRUE, B{row}:E{row}), "")'
    apply_styles(ws_meta_adset, META_ADSET_COLS, META_ADSET_VALS, formula_col_idx=6, formula_template=meta_adset_formula)

    # ---------------------------------------------------------
    # 3. META AD NOMENCLATURE
    # [INT/INF]_[Ad Format]_[Product]_[Language]_[Creator/Model Name]_[Narrative]_[Hook/Variant]
    # ---------------------------------------------------------
    ws_meta_ad = wb.create_sheet("Meta Ad")
    META_AD_COLS = [
        ("S.No.", 8),
        ("INT/INF", 15),
        ("Ad Format", 20),
        ("Product", 25),
        ("Language", 15),
        ("Creator/Model Name", 25),
        ("Narrative", 25),
        ("Hook/Variant (Type Here)", 30),
        ("Generated Ad Name", 90)
    ]
    META_AD_VALS = {
        "INT/INF": '"INF,Internal_UGC,Internal_Brand"',
        "Ad Format": '"Reel,Static,Carousel,Video,Boomerang,GIF"',
        "Language": '"English,Hindi,Hinglish,Marathi"',
        "Narrative": '"Trust/Safe,Offer-led,Feature-led,Problem_Solution,Founder_Story"'
    }
    meta_ad_formula = '=IF(COUNTA(B{row}:H{row})>0, TEXTJOIN("_", TRUE, B{row}:H{row}), "")'
    apply_styles(ws_meta_ad, META_AD_COLS, META_AD_VALS, formula_col_idx=9, formula_template=meta_ad_formula)

    # ---------------------------------------------------------
    # 4. GOOGLE CAMPAIGN NOMENCLATURE
    # [Platform]_[Funnel Stage]_[Objective]_[Product/Category]_[Promo/Evergreen]_[Geo]
    # ---------------------------------------------------------
    ws_google_camp = wb.create_sheet("Google Campaign")
    GOOGLE_CAMP_COLS = [
        ("S.No.", 8),
        ("Platform", 15),
        ("Funnel Stage", 20),
        ("Objective", 20),
        ("Product/Category", 25),
        ("Promo/Evergreen", 20),
        ("Geo", 20),
        ("Generated Campaign Name", 70)
    ]
    GOOGLE_CAMP_VALS = {
        "Platform": '"Google"',
        "Funnel Stage": '"Search,PMax,Shopping,Display,YouTube,Demand_Gen"',
        "Objective": '"Sales,Leads,Traffic,Awareness"',
        "Product/Category": '"Haircare,Skincare,Bodycare,Supplements,Kits,All Products"',
        "Promo/Evergreen": '"Brand,Non-Brand,Competitor,Generic,Evergreen,Promo"',
        "Geo": '"PanIndia,Tier1,Tier2,Metro,US"'
    }
    google_camp_formula = '=IF(COUNTA(B{row}:G{row})>0, TEXTJOIN("_", TRUE, B{row}:G{row}), "")'
    apply_styles(ws_google_camp, GOOGLE_CAMP_COLS, GOOGLE_CAMP_VALS, formula_col_idx=8, formula_template=google_camp_formula)

    # ---------------------------------------------------------
    # 5. GOOGLE AD GROUP NOMENCLATURE
    # [Audience Type]_[Audience Details]_[Demographics]_[Placement]
    # ---------------------------------------------------------
    ws_google_adgroup = wb.create_sheet("Google Ad Group")
    GOOGLE_ADGROUP_COLS = [
        ("S.No.", 8),
        ("Audience Type", 25),
        ("Audience Details (e.g., Best_Hair_Serum, Cart_Abandoners)", 45),
        ("Demographics", 20),
        ("Placement/Format", 20),
        ("Generated Ad Group Name", 70)
    ]
    GOOGLE_ADGROUP_VALS = {
        "Audience Type": '"Exact_Match,Phrase_Match,Broad_Match,Audience_Signals,Placement_Targeted,Retargeting"',
        "Demographics": '"18-35,25-44,All,N/A"',
        "Placement/Format": '"Search,Display,YouTube,Shopping,PMax_Asset"'
    }
    google_adgroup_formula = '=IF(COUNTA(B{row}:E{row})>0, TEXTJOIN("_", TRUE, B{row}:E{row}), "")'
    apply_styles(ws_google_adgroup, GOOGLE_ADGROUP_COLS, GOOGLE_ADGROUP_VALS, formula_col_idx=6, formula_template=google_adgroup_formula)

    # ---------------------------------------------------------
    # 6. GOOGLE AD NOMENCLATURE
    # [Ad Format]_[Product]_[Language]_[Narrative]_[Asset/Concept]
    # ---------------------------------------------------------
    ws_google_ad = wb.create_sheet("Google Ad")
    GOOGLE_AD_COLS = [
        ("S.No.", 8),
        ("Ad Format", 20),
        ("Product", 25),
        ("Language", 15),
        ("Narrative", 25),
        ("Asset/Concept (Type)", 30),
        ("Generated Ad Name", 80)
    ]
    GOOGLE_AD_VALS = {
        "Ad Format": '"RSA,RDA,Video_Action,Bumper,Skippable,Asset_Group"',
        "Language": '"English,Hindi,Hinglish,Marathi"',
        "Narrative": '"Trust/Safe,Offer-led,Feature-led,Problem_Solution,Founder_Story"'
    }
    google_ad_formula = '=IF(COUNTA(B{row}:F{row})>0, TEXTJOIN("_", TRUE, B{row}:F{row}), "")'
    apply_styles(ws_google_ad, GOOGLE_AD_COLS, GOOGLE_AD_VALS, formula_col_idx=7, formula_template=google_ad_formula)
    
    # Save the Workbook
    output_path = os.path.join(os.getcwd(), FILE_NAME)
    wb.save(output_path)
    print(f"Success! Saved to {output_path}")

if __name__ == "__main__":
    create_dashboard()
