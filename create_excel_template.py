import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_NAME = "D2C_Marketing_Dashboard_Template.xlsx"

# Colors
COLOR_SLATE_800 = "1E293B"
COLOR_SLATE_100 = "F1F5F9"
COLOR_WHITE     = "FFFFFF"

# Styles
font_bold = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
fill_header = PatternFill(start_color=COLOR_SLATE_800, end_color=COLOR_SLATE_800, fill_type='solid')

# Borders
BORDER_THIN = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

def create_template():
    print("Generating Empty Excel Template with active auto-tagging and rate calculation formulas...")
    
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. META_DATA SHEET
    ws_meta = wb.create_sheet(title="META_DATA")
    meta_headers = [
        'Day', 'Account Name', 'Campaign name', 'Ad set name', 'Ad name', 'Currency', 'Amount spent (INR)', 
        'Impressions', 'Link clicks', 'Landing page views', 'Add To Cart', 'Purchases', 'Purchases conversion value', 
        '3-second video plays', 'Video plays at 25%', 'Video plays at 50%', 'Video plays at 75%', 'Video plays at 95%', 'Video plays at 100%', 
        'Website URL', 'Ad set ID', 'Ad ID', 
        'Product', 'Ad Type', 'Type of Ad', 'Influencer Name', 'Narrative', 'Language',
        'ROAS', 'Hook Rate %', 'Hold Rate %'
    ]
    ws_meta.append(meta_headers)
    
    # Pre-fill row formulas for rows 2 to 5000
    for row in range(2, 5001):
        comb = f'C{row}&"|"&D{row}&"|"&E{row}'
        
        # Build product mapping formula
        rules = [
            ('new internal', 'Hair Gummies'),
            ('new_internal', 'Hair Gummies'),
            ('folli', 'Hair Growth Roll On Folli Advanced'),
            ('stretch mark', 'Anti Stretch Mark Roll On'),
            ('stretch_mark', 'Anti Stretch Mark Roll On'),
            ('30', '30 Urea Foot Roll On'),
            ('urea kit', 'Urea Skin & Heel Repair Kit'),
            ('heel repair kit', 'Urea Skin & Heel Repair Kit'),
            ('urea skin', 'Urea Skin & Heel Repair Kit'),
            ('20% urea lotion', '20% UL'),
            ('20 urea lotion', '20% UL'),
            ('20_urea_lotion', '20% UL'),
            ('20% ul', '20% UL'),
            ('20 ul', '20% UL'),
            ('20_ul', '20% UL'),
            ('10% urea lotion', 'UL'),
            ('10 urea lotion', 'UL'),
            ('10_urea_lotion', 'UL'),
            ('10% ul', 'UL'),
            ('10 ul', 'UL'),
            ('10_ul', 'UL'),
            (' ul ', 'UL'),
            ('ceramide', '1% Salicylic + Ceramide Bodywash'),
            ('triple milk', 'Triple Milk Body Wash'),
            ('triple_milk', 'Triple Milk Body Wash'),
            ('lactic acid', '5% Lactic Acid Lotion'),
            ('lactic_acid', '5% Lactic Acid Lotion'),
            ('lactic lotion', '5% Lactic Acid Lotion'),
            ('lactic_lotion', '5% Lactic Acid Lotion'),
            ('niacinamide', '10% Niacinamide Lotion'),
            ('6% underarm pigmentation kit', 'Advanced Underarm Pigmentation Kit'),
            ('6% underarm pigmentation', 'Advanced Underarm Pigmentation Kit'),
            ('6% underarm', '6% AHA BHA Underarm Roll On'),
            ('6% roll', '6% AHA BHA Underarm Roll On'),
            ('6% uaro', '6% AHA BHA Underarm Roll On'),
            ('6%_uaro', '6% AHA BHA Underarm Roll On'),
            ('4% underarm roll', '4% AHA BHA Underarm Roll On'),
            ('4% & 6% underarm roll', '4% AHA BHA Underarm Roll On'),
            ('4% aha bha underarm', '4% AHA BHA Underarm Roll On'),
            ('underarm_pigmentation_pack', 'Advanced Underarm Pigmentation Kit'),
            ('pigmentation_pack', 'Advanced Underarm Pigmentation Kit'),
            ('pigmentation pack', 'Advanced Underarm Pigmentation Kit'),
            ('advanced underarm pigmentation', 'Advanced Underarm Pigmentation Kit'),
            ('underarm pigmentation kit', 'Advanced Underarm Pigmentation Kit'),
            ('upk', 'Advanced Underarm Pigmentation Kit'),
            ('underarm roll', '4% AHA BHA Underarm Roll On'),
            ('underarm_roll', '4% AHA BHA Underarm Roll On'),
            ('underarm_roll_on', '4% AHA BHA Underarm Roll On'),
            ('uaro', '4% AHA BHA Underarm Roll On'),
            ('urea roll', '20% Urea Foot Roll On'),
            ('urea_roll', '20% Urea Foot Roll On'),
            ('urea', '20% Urea Foot Roll On'),
            ('calcium gummies', 'Calcium Gummies'),
            ('calcium_gummies', 'Calcium Gummies'),
            ('calcium', 'Calcium Gummies'),
            ('magnesium lotion', 'Magnesium Lotion'),
            ('magnesium_lotion', 'Magnesium Lotion'),
            ('magnesium gummies', 'Magnesium Glycinate Gummies'),
            ('magnesium_gummies', 'Magnesium Glycinate Gummies'),
            ('magnesium', 'Magnesium Glycinate Gummies'),
            ('glutathione', 'Glutathione Gummies'),
            ('shilajit', 'Shilajit Gummies'),
            ('revised mrp', 'Hair Gummies (549)'),
            ('revised_mrp', 'Hair Gummies (549)'),
            ('postpartum', 'HG - PP'),
            ('post_partum', 'HG - PP'),
            ('_pp_', 'HG - PP'),
            ('_pp |', 'HG - PP'),
            ('hair growth pack', 'AHGP (Roll On)'),
            ('hair_growth_pack', 'AHGP (Roll On)'),
            ('ahgp', 'AHGP (Roll On)'),
            ('hair gummies', 'Hair Gummies'),
            ('hair_gummies', 'Hair Gummies'),
            ('hair growth roll', 'Hair Growth Roll On'),
            ('hair_growth_roll', 'Hair Growth Roll On'),
            ('advanced hair growth', 'Hair Growth Roll On'),
            ('ahgs', 'Hair Growth Roll On'),
            ('keto shampoo', '1% Keto Shampoo'),
            ('ketoconazole', '1% Keto Shampoo'),
            ('keto_shampoo', '1% Keto Shampoo'),
            ('glycolic stick', 'Glycolic Stick'),
            ('glycolic_stick', 'Glycolic Stick'),
            ('glycolic', 'Glycolic Stick'),
            ('hydration kit', 'Hydration Kit'),
            ('hydration_kit', 'Hydration Kit'),
            ('5% aha bha exfoliating body wash', '5% AHA BHA Exfoliating Body Wash'),
            ('5% aha bha bodywash', '5% AHA BHA Exfoliating Body Wash'),
            ('5% bodywash', '5% AHA BHA Exfoliating Body Wash'),
            ('5%_aha_bha_bodywash', '5% AHA BHA Exfoliating Body Wash'),
            ('exfoliating body wash', '5% AHA BHA Exfoliating Body Wash'),
            ('1% sa bodywash adset', '1% Salicylic Acid Body Wash - 250 mL'),
            ('body acne pack', 'Body Acne Pack'),
            ('body_acne_pack', 'Body Acne Pack'),
            ('sa bodywash', '1% Salicylic Acid Body Wash'),
            ('sa_bodywash', '1% Salicylic Acid Body Wash'),
            ('salicylic acid body wash', '1% Salicylic Acid Body Wash'),
            ('sa body wash', '1% Salicylic Acid Body Wash'),
            ('wintercare', 'Wintercare'),
            ('winter care', 'Winter Care'),
            ('winter_care', 'Winter Care'),
            ('strawberry', 'Straw'),
            ('mehr_hair_growth', 'HGP'),
            ('hgp', 'HGP'),
            ('all_body_products', 'Body Category'),
            ('all_body', 'Body Category'),
            ('all_hair_products', 'Hair Category'),
            ('all_hair', 'Hair Category'),
            ('skin_category', 'Body Category'),
            ('body_care', 'Body Category'),
            ('acne_treatment', 'Body Category'),
            ('sulphur', 'Body Category'),
            ('body care', 'Body Category'),
            ('category page', 'IF(ISNUMBER(SEARCH("body",' + comb + ')),"Body Category","Hair Category")'),
            ('category_page', 'IF(ISNUMBER(SEARCH("body",' + comb + ')),"Body Category","Hair Category")'),
            ('all_kit', 'Kits'),
            ('kit_page', 'Kits'),
            ('all kit', 'Kits'),
            ('kit page', 'Kits'),
            ('shop all', 'Kits'),
            ('pack_na', 'Kits'),
            ('hgro', 'Kits'),
            ('asc_all_kit_page', 'Kits'),
            ('all_kit_page', 'Kits')
        ]
        
        formula = '"Other Product"'
        for sub, res in reversed(rules):
            if res.startswith('IF('):
                formula = f'IF(ISNUMBER(SEARCH("{sub}",{comb})),{res},{formula})'
            else:
                formula = f'IF(ISNUMBER(SEARCH("{sub}",{comb})),"{res}",{formula})'
        f_prod = f'=IF(E{row}="","", {formula})'
        
        f_adtype = (
            f'=IF(E{row}="","",IF(OR(ISNUMBER(SEARCH("influencer",E{row})),ISNUMBER(SEARCH("infl",E{row})),ISNUMBER(SEARCH("collab",E{row})),ISNUMBER(SEARCH("creator",E{row})),ISNUMBER(SEARCH("inf",E{row}))),'
            f'"Influencer","Internal UGC"))'
        )
        
        f_format = (
            f'=IF(E{row}="","",IF(ISNUMBER(SEARCH("carousel",E{row})),"Carousel",'
            f'IF(OR(ISNUMBER(SEARCH("static",E{row})),ISNUMBER(SEARCH("image",E{row}))),"Static Creative",'
            f'IF(OR(ISNUMBER(SEARCH("voice",E{row})),ISNUMBER(SEARCH("reel",E{row})),ISNUMBER(SEARCH("video",E{row})),ISNUMBER(SEARCH("shorts",E{row}))),'
            f'IF(X{row}="Influencer","Reel","Video Ad"),"Video Ad"))))'
        )
        
        inf_rules = [
            ('affluence', 'Affluence'),
            ('bhavi', 'Bhavi'),
            ('pooja', 'Pooja'),
            ('esha shetty', 'Esha Shetty'),
            ('esha_shetty', 'Esha Shetty'),
            ('surabhi', 'Surabhi'),
            ('shilpa', 'Shilpa'),
            ('timsy', 'Timsy'),
            ('chitwan', 'Chitwan Garg'),
            ('mehvi', 'Mehvi Thapa'),
            ('suhana', 'Suhana Grover'),
            ('swati chauhan', 'Swati Chauhan'),
            ('swati_chauhan', 'Swati Chauhan'),
            ('swati', 'Swati'),
            ('khushboo', 'Khushboo'),
            ('buvana', 'Buvana'),
            ('deepika rani', 'Deepika Rani'),
            ('deepika_rani', 'Deepika Rani'),
            ('deepali', 'Deepali'),
            ('shamli', 'Shamli'),
            ('nikkita', 'Nikkitha'),
            ('nikitta', 'Nikkitha'),
            ('nikkitha', 'Nikkitha'),
            ('sarita', 'Sarita'),
            ('sumegha', 'Sumegha'),
            ('jiya', 'Jiya'),
            ('rini', 'Rini'),
            ('yogita', 'Yogita'),
            ('jyoti', 'Jyoti'),
            ('vaishnavi', 'Vaishnavi'),
            ('shivanti', 'Dr. Shivanti'),
            ('amrita', 'Amrita'),
            ('pria', 'Pria'),
            ('mahi', 'Mahi'),
            ('ridhi khanna', 'Riddhi Khanna'),
            ('ridhi_khanna', 'Riddhi Khanna'),
            ('riddhi khanna', 'Riddhi Khanna'),
            ('riddhi_khanna', 'Riddhi Khanna'),
            ('pritha', 'Pritha Khot')
        ]
        inf_formula = f'IF(X{row}="Influencer", "Influencer", "Internal")'
        for sub, res in reversed(inf_rules):
            inf_formula = f'IF(ISNUMBER(SEARCH("{sub}",{comb})),"{res}",{inf_formula})'
        f_influencer = f'=IF(E{row}="","", {inf_formula})'
        
        narr_rules = [
            ('discount', 'Offer-led'),
            ('sale', 'Offer-led'),
            ('coupon', 'Offer-led'),
            ('price', 'Offer-led'),
            ('offer', 'Offer-led'),
            ('off', 'Offer-led'),
            ('benefit', 'Feature-led'),
            ('feature', 'Feature-led'),
            ('fabric', 'Feature-led'),
            ('quality', 'Feature-led'),
            ('comfort', 'Feature-led'),
            ('unboxing', 'Social Proof/UGC'),
            ('review', 'Social Proof/UGC'),
            ('testimonial', 'Social Proof/UGC'),
            ('ugc', 'Social Proof/UGC'),
            ('founder', 'Founder Story'),
            ('story', 'Founder Story'),
            ('why', 'Founder Story'),
            ('journey', 'Founder Story'),
            ('trust', 'Trust/Safe Ingredients'),
            ('safe', 'Trust/Safe Ingredients'),
            ('ingredients', 'Trust/Safe Ingredients'),
            ('all_in_one', 'All In One Bodywash Detan'),
            ('all in one', 'All In One Bodywash Detan'),
            ('detan', 'All In One Bodywash Detan')
        ]
        narr_formula = '"Brand Story"'
        for sub, res in reversed(narr_rules):
            narr_formula = f'IF(ISNUMBER(SEARCH("{sub}",{comb})),"{res}",{narr_formula})'
        f_narrative = f'=IF(E{row}="","", {narr_formula})'
        
        f_lang = (
            f'=IF(E{row}="","",IF(ISNUMBER(SEARCH("hinglish",E{row})),"Hinglish",'
            f'IF(ISNUMBER(SEARCH("hindi",E{row})),"Hindi",'
            f'IF(ISNUMBER(SEARCH("marathi",E{row})),"Marathi",'
            f'IF(ISNUMBER(SEARCH("tamil",E{row})),"Tamil",'
            f'IF(ISNUMBER(SEARCH("telugu",E{row})),"Telugu",'
            f'IF(ISNUMBER(SEARCH("kannada",E{row})),"Kannada",'
            f'IF(ISNUMBER(SEARCH("malayalam",E{row})),"Malayalam",'
            f'IF(ISNUMBER(SEARCH("punjabi",E{row})),"Punjabi","English"))))))))'
        )
        
        f_roas = f'=IF(E{row}="","",IFERROR(M{row}/G{row},0))'
        f_hook = f'=IF(E{row}="","",IFERROR(N{row}/H{row},0))'
        f_hold = f'=IF(E{row}="","",IFERROR(O{row}/N{row},0))'
        
        # We write empty values for columns A-V and the formulas for W-AE
        ws_meta.append([''] * 22 + [f_prod, f_adtype, f_format, f_influencer, f_narrative, f_lang, f_roas, f_hook, f_hold])
        
    # Formatting META_DATA sheet
    ws_meta.views.sheetView[0].showGridLines = True
    for col in range(1, len(meta_headers) + 1):
        cell = ws_meta.cell(row=1, column=col)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center')
        
    for row in range(2, 5001):
        ws_meta.cell(row=row, column=7).number_format = '₹#,##0.00'   # Spend
        ws_meta.cell(row=row, column=8).number_format = '#,##0'       # Imp
        ws_meta.cell(row=row, column=9).number_format = '#,##0'       # Clicks
        ws_meta.cell(row=row, column=10).number_format = '#,##0'      # LPV
        ws_meta.cell(row=row, column=11).number_format = '#,##0'      # ATC
        ws_meta.cell(row=row, column=12).number_format = '#,##0'      # Purchases
        ws_meta.cell(row=row, column=13).number_format = '₹#,##0.00'  # Purchases Value
        ws_meta.cell(row=row, column=29).number_format = '0.00"x"'    # ROAS Formula
        ws_meta.cell(row=row, column=30).number_format = '0.00%'      # Hook Formula
        ws_meta.cell(row=row, column=31).number_format = '0.00%'      # Hold Formula
        
    # Auto-adjust column widths
    for col in ws_meta.columns:
        col_letter = get_column_letter(col[0].column)
        ws_meta.column_dimensions[col_letter].width = 16
        
    # 2. GOOGLE_DATA SHEET
    ws_google = wb.create_sheet(title="GOOGLE_DATA")
    google_headers = [
        'Date','Campaign','Ad Group','Ad Name',
        'Spend','Impressions','Clicks','Link Clicks','Landing Page Views',
        'Conversions','Revenue','Add to Cart','CTR %','CPC','ROAS',
        'Quality Score','Conv Rate %','Hook Rate %','Hold Rate %',
        'Narrative','Type of Ad (Static/Reel)','Language of the Ad','Name of the Influencer','Product Name','Category'
    ]
    ws_google.append(google_headers)
    ws_google.views.sheetView[0].showGridLines = True
    for col in range(1, len(google_headers) + 1):
        cell = ws_google.cell(row=1, column=col)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center')
        col_letter = get_column_letter(col)
        ws_google.column_dimensions[col_letter].width = 16
        
    # 3. SALES_DATA SHEET
    ws_sales = wb.create_sheet(title="SALES_DATA")
    sales_headers = ['Order ID', 'Customer ID', 'Order Date', 'Revenue', 'Product', 'Platform', 'Cohort Month']
    ws_sales.append(sales_headers)
    ws_sales.views.sheetView[0].showGridLines = True
    for col in range(1, len(sales_headers) + 1):
        cell = ws_sales.cell(row=1, column=col)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center')
        col_letter = get_column_letter(col)
        ws_sales.column_dimensions[col_letter].width = 16
        
    # 4. PRODUCT_MASTER SHEET
    ws_prod = wb.create_sheet(title="PRODUCT_MASTER")
    prod_headers = ['Product', 'Price', 'Cost of Goods', 'Margin', 'Target CPA']
    ws_prod.append(prod_headers)
    ws_prod.views.sheetView[0].showGridLines = True
    for col in range(1, len(prod_headers) + 1):
        cell = ws_prod.cell(row=1, column=col)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center')
        col_letter = get_column_letter(col)
        ws_prod.column_dimensions[col_letter].width = 30
        
    # Pre-populate default product rows based on user's real inventory
    product_master = [
        ['Advanced Hair Growth Roll On', 799, 280, 0.65, 300],
        ['5% AHA BHA Bodywash', 499, 175, 0.65, 180],
        ['All Body Products', 1499, 525, 0.65, 500],
        ['Underarm Pigmentation Pack', 999, 350, 0.65, 350],
        ['Underarm Roll On', 499, 175, 0.65, 180],
        ['Biotin Hair Gummies', 599, 210, 0.65, 200],
        ['Hair Growth Serum', 699, 245, 0.65, 250],
        ['De-Tan Sunscreen', 399, 140, 0.65, 150]
    ]
    for p in product_master:
        ws_prod.append(p)
    for row in range(2, ws_prod.max_row + 1):
        ws_prod.cell(row=row, column=2).number_format = '₹#,##0.00'
        ws_prod.cell(row=row, column=3).number_format = '₹#,##0.00'
        ws_prod.cell(row=row, column=4).number_format = '0.0%'
        ws_prod.cell(row=row, column=5).number_format = '₹#,##0.00'
        
    # 5. CONFIG SHEET
    ws_cfg = wb.create_sheet(title="CONFIG")
    cfg_headers = ['Setting Name', 'Value', 'Description']
    ws_cfg.append(cfg_headers)
    ws_cfg.views.sheetView[0].showGridLines = True
    for col in range(1, len(cfg_headers) + 1):
        cell = ws_cfg.cell(row=1, column=col)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center')
        col_letter = get_column_letter(col)
        ws_cfg.column_dimensions[col_letter].width = 25
        
    config_rows = [
        ['ROAS', 1.5, 'Minimum acceptable Target ROAS'],
        ['Scale Threshold', 4.0, 'ROAS threshold to recommend scaling an ad'],
        ['CTR', 2.5, 'Click-Through Rate benchmark (%)'],
        ['Conv Rate', 2.0, 'Conversion Rate benchmark (%)'],
        ['Hook Rate', 30.0, 'Video Hook Rate benchmark (%)'],
        ['Hold Rate', 20.0, 'Video Hold Rate benchmark (%)'],
        ['Target LPV Rate', 70.0, 'Target Landing Page Views Rate (%)'],
        ['Target ATC Rate', 5.0, 'Target Add To Cart Rate (%)']
    ]
    for c in config_rows:
        ws_cfg.append(c)
        
    # Save Workbook
    wb.save(FILE_NAME)
    print(f"Template successfully generated: {FILE_NAME}")
    
    # Also save copy to Desktop
    desktop_path = "/Users/balbaasaur/Desktop/D2C_Marketing_Dashboard_Template.xlsx"
    try:
        wb.save(desktop_path)
        print(f"Copied template to Desktop: {desktop_path}")
    except Exception as e:
        print(f"Failed to copy template to Desktop: {e}")

if __name__ == "__main__":
    create_template()
